from openpyxl import *
from tkinter import *
import re
from tkinter import messagebox
from tkinter import ttk

wb = load_workbook(r"D:/python 2111812 (1)/lab04/hehe.xlsx")
sheet = wb.active

def excel():
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 40
    sheet.column_dimensions['G'].width = 50
    sheet.column_dimensions['H'].width = 30

    sheet.cell(row=1, column=1).value = "Mã số sinh viên"
    sheet.cell(row=1, column=2).value = "Họ tên"
    sheet.cell(row=1, column=3).value = "Ngày sinh"
    sheet.cell(row=1, column=4).value = "Email"
    sheet.cell(row=1, column=5).value = "Số điện thoại"
    sheet.cell(row=1, column=6).value = "Học kỳ"
    sheet.cell(row=1, column=7).value = "Năm học"

def clear():
    num_field.delete(0, END)
    name_field.delete(0, END)
    date_field.delete(0, END)
    email_field.delete(0, END)
    phone_num_no_field.delete(0, END)
    sem_field.delete(0, END)
    year_field.delete(0, END)

def validate_student_id(event):
    student_id = num_field.get()
    if not student_id.isdigit() or len(student_id) != 7:
        messagebox.showerror("Error", "Vui lòng chỉ nhập số vào ô mã số sinh viên và phải có 7 kí tự!")
        
def focus1(event):
    name_field.focus_set()

 
def email_HopLe(email):
    pattern = r'^[\w\.-]+@[\w\.-]+\.\w+$'  #bieu thuc chinh quy 
    if not re.match(pattern, email):
        messagebox.showerror("Error","Email không hợp lệ! Vui lòng nhập lại!")
        return False
    return True

def validate_email():
    email = email_field.get()
    if not email_HopLe(email):
        return 
        
def validate_phone_number():
    phone = num_field.get()
    if not phone.isdigit() or len(phone) != 10:
        messagebox.showerror("Error", "Vui lòng chỉ nhập số vào ô số điện thoại và phải có 10 chữ số!")  
        
def semester_HopLe(semester):
    semester = sem_field.get()
    if semester not in ['1', '2', '3']:
        messagebox.showerror("Loi", "Học kỳ không hợp lệ!")
        return False
    return True


def validate_semester(event):
    semester = sem_field.get()
    if not semester_HopLe(semester):
        messagebox.showerror("Error", "Học kỳ không hợp lệ!")
        
def date_of_birth_HopLe(date_of_birth):
    pattern = r'^\d{2}/\d{2}/\d{4}$'
    if not re.match(pattern, date_of_birth):
        messagebox.showerror("Loi", "Ngày tháng năm sinh không hợp lệ!")
        return False
    return True

        
def validate_date_of_birth(event):
    date_of_birth = date_field.get()
    if not date_of_birth_HopLe(date_of_birth):
        messagebox.showerror("Error", "Ngày sinh không hợp lệ!")
        

    
    

def insert():
    student_id = num_field.get()
    name = name_field.get()
    date_of_birth = date_field.get()
    email = email_field.get()
    phone = phone_num_no_field.get()
    semester = sem_field.get()
    year = year_field.get()
    
    if (student_id == "" or name == "" or date_of_birth == "" or email == "" or phone == "" or semester == "" or year == ""):
        messagebox.showerror("Error", "Vui lòng điền đầy đủ thông tin!")
        return
    
    if not student_id.isdigit() or len(student_id) != 7:
        messagebox.showerror("Error", "Vui lòng chỉ nhập số vào ô mã số sinh viên và phải có 7 kí tự!")
        return

    if not email_HopLe(email):
        return

    if not phone.isdigit() or len(phone) != 10:
        messagebox.showerror("Error", "Vui lòng chỉ nhập số vào ô số điện thoại và phải có 10 kí tự!")
        return
        
    if not semester_HopLe(semester):
        return
        
    if not date_of_birth_HopLe(date_of_birth):
        return
        
    '''if not namhoc_HopLe(year):
        return'''

    current_row = sheet.max_row
    sheet.cell(row=current_row + 1, column=1).value = student_id
    sheet.cell(row=current_row + 1, column=2).value = name
    sheet.cell(row=current_row + 1, column=3).value = date_of_birth
    sheet.cell(row=current_row + 1, column=4).value = email
    sheet.cell(row=current_row + 1, column=5).value = phone
    sheet.cell(row=current_row + 1, column=6).value = semester
    sheet.cell(row=current_row + 1, column=7).value = year

    wb.save("D:\python 2111812 (1)\lab04\hehe.xlsx")

    clear()

    messagebox.showinfo("Thông báo", "Đăng ký thành công!")
    
if __name__ == "__main__":
    

    root = Tk()

    root.configure(background='light green')
    
    root.title("THÔNG TIN ĐĂNG KÝ HỌC PHẦN")

    root.geometry("500x300")
 
    excel()
 
    heading = Label(root, text="THÔNG TIN ĐĂNG KÝ HỌC PHẦN", bg="light green", fg="RED", font=(20))

    num = Label(root, text="Mã số sinh viên", bg="light green")
    
    name = Label(root, text="Họ tên", bg="light green")
 
    date = Label(root, text="Ngày sinh", bg="light green")
 
    email = Label(root, text="Email", bg="light green")
 
    phone_num_no = Label(root, text="Số điện thoại", bg="light green")
 
    sem = Label(root, text="Học kỳ", bg="light green")
 
    year = Label(root, text="Năm học", bg="light green")
    
    choose = Label(root, text="Chọn môn học", bg="light green")
     
 

    heading.grid(row=0, column=1, columnspan=2)
    num.grid(row=1, column=0, sticky="w", padx=(0,10), )
    name.grid(row=2, column=0,sticky="w")
    date.grid(row=3, column=0,sticky="w")
    email.grid(row=4, column=0,sticky="w")
    phone_num_no.grid(row=5, column=0,sticky="w")
    sem.grid(row=6, column=0,sticky="w")
    year.grid(row=7, column=0,sticky="w")
    year_var = StringVar(root)
    choose.grid(row=8, column=0,sticky="w")

 
    num_field = Entry(root)
    name_field = Entry(root)
    date_field = Entry(root)
    email_field = Entry(root)
    phone_num_no_field = Entry(root)
    sem_field = Entry(root)
    year_field = Entry(root)
    year_option = ["2022-2023", "2023-2024", "2024-2025"]
    year_var = StringVar(root)
    year_var.set(year_option[0])
 
    num_field.bind("<Return>", validate_student_id)                 #ẩn chức năng
    
    name_field.bind("<Return>", focus1)

    #date_field.bind("<Return>", validate_date_of_birth)

    email_field.bind("<Return>", validate_email)

    phone_num_no_field.bind("<Return>", validate_phone_number)
 
    sem.bind("<Return>", validate_semester)
    
    #year.bind("<Return>", validate_year)
    
    #choose.bind("<Return>", focus8)
    



    num_field.grid(row=1, column=1, ipadx="100", columnspan=2)
    name_field.grid(row=2, column=1, ipadx="100", columnspan=2)
    date_field.grid(row=3, column=1, ipadx="100", columnspan=2)
    email_field.grid(row=4, column=1, ipadx="100", columnspan=2)
    phone_num_no_field.grid(row=5, column=1, ipadx="100", columnspan=2)
    sem_field.grid(row=6, column=1, ipadx="100", columnspan=2)
    year_field = ttk.Combobox(root, textvariable=year_var, values=year_option)
    year_field.grid(row=7, column=1, ipadx="90", columnspan=2)

    
    excel()
    subject1 = Checkbutton(root, text="Lập trình Python", bg = "light green", command=Checkbutton)
    subject1.grid(row=8, column=1, sticky="w",columnspan=1)
    
    subject2 = Checkbutton(root, text="Lập trình Java", bg = "light green", command=Checkbutton)
    subject2.grid(row=8, column=2, sticky="w",columnspan=1)
    
    subject3 = Checkbutton(root, text="Công nghệ phần mềm", bg = "light green", command=Checkbutton)
    subject3.grid(row=9, column=1, sticky="w",columnspan=1)
    
    subject4 = Checkbutton(root, text="Phát triển ứng dụng web", bg = "light green", command=Checkbutton)
    subject4.grid(row=9, column=2, sticky="w",columnspan=2)
 
    submit = Button(root, text="Submit", bg="lime", command=insert)
    submit.grid(row=10, column=1)
    
    exit = Button(root, text="Thoát", bg="lime", command=exit)
    exit.grid(row = 10, column=2)
 
    root.mainloop()
root = Tk()

